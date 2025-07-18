import os
import json
import time
import anthropic
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

client = anthropic.Anthropic(
    api_key=os.getenv("CLAUDE_API_KEY")
)

TEXT_DIR = "../tenders"
OUTPUT_EXCEL = "../tenders/claude_extracted.xlsx" 
MAX_FILES = 3
MAX_TOKENS = 8000 

# Enhanced columns based on stakeholder requirements
COLUMNS = [
    "Title", "Issuer", "Deadline", "Budget", "Location", 
    "Project Type", "Required Documents", "PC AVK5 Required",
    "Technical Specifications", "Payment Terms", "Resource Requirements",
    "Timeline Feasibility", "Profitability Assessment", "Filename"
]

data = []
def build_tender_text(tender_json):
    title = tender_json.get("title", "")
    description = tender_json.get("description", "")
    issuer = tender_json.get("procuringEntity", {}).get("name", "")
    address = tender_json.get("procuringEntity", {}).get("address", {})
    location = f"{address.get('locality', '')}, {address.get('region', '')}".strip(", ")
    budget = tender_json.get("value", {}).get("amount", "N/A")
    currency = tender_json.get("value", {}).get("currency", "UAH")
    deadline = tender_json.get("tenderPeriod", {}).get("endDate", "Not specified")

    items = tender_json.get("items", [])
    item_descriptions = [
        f"- {item.get('description', '')} ({item.get('classification', {}).get('description', '')})"
        for item in items
    ]

    tech_specs = []
    for criterion in tender_json.get("criteria", []):
        for group in criterion.get("requirementGroups", []):
            for req in group.get("requirements", []):
                title = req.get("title", "")
                expected = req.get("expectedValues", []) or [req.get("expectedValue", "")]
                if title:
                    tech_specs.append(f"{title}: {', '.join(str(v) for v in expected if v)}")

    return f"""
Tender Title: {title}
Issuer: {issuer}
Location: {location}
Budget: {budget} {currency}
Deadline: {deadline}

Goods/Services:
{chr(10).join(item_descriptions)}

Description:
{description}

Technical Requirements:
{chr(10).join(tech_specs)}
""".strip()

def ask_claude(text):
    prompt = f"""
You are an expert in Ukrainian public procurement tenders. Analyze the tender text and extract the following information from the file:

1. Basic Information:
   - Title or Project Name
   - Issuer or Client
   - Submission Deadline
   - Estimated Budget (with currency)
   - Location (City, Region)
   - Project Type/Scope

2. Critical Requirements:
   - List ALL required documents (comma-separated)
   - Does this tender require PC AVK5 cost estimates? (true/false)
   - Key technical specifications (summarize key requirements)

3. Financial & Legal:
   - Payment terms and schedule
   - References to Ukrainian laws/regulations (list)

4. Viability Analysis:
   - Resource requirements (equipment, personnel, etc.)
   - Timeline feasibility assessment (adequate/risky/inadequate)
   - Profitability assessment (high/medium/low)

Return the result STRICTLY in JSON format with these keys:
{{
  "title": "...",
  "issuer": "...",
  "deadline": "...",
  "budget": "...",
  "location": "...",
  "project_type": "...",
  "required_documents": ["doc1", "doc2", ...],
  "avk5_required": true/false,
  "technical_specs": "...",
  "payment_terms": "...",
  "resource_requirements": "...",
  "timeline_feasibility": "...",
  "profitability": "..."
}}

Tender Text:
\"\"\"
{text}
\"\"\"
"""
    response = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=1024,
        temperature=0.0,
        system="You are a procurement specialist analyzing Ukrainian tenders. Focus on PC AVK5 compliance and document requirements.",
        messages=[{"role": "user", "content": prompt}]
    )
    
    if response.content:
        return ''.join(block.text for block in response.content if hasattr(block, 'text')) #type: ignore
    return ""

def format_excel(file_path):
    wb = Workbook()
    ws = wb.active
    
    # Header formatting
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths based on content type
    col_widths = [40, 30, 15, 15, 20, 25, 40, 15, 50, 30, 40, 20, 20, 30]
    
    for col_num, (column_title, width) in enumerate(zip(COLUMNS, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)  # type: ignore
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width  # type: ignore
    
    return wb, ws

def save_to_excel(ws, row_data, row_counter):
    for col_num, col_name in enumerate(COLUMNS, 1):
        value = row_data.get(col_name, "N/A")
        
        # Handle list types
        if isinstance(value, list):
            value = ", ".join(value)
        
        cell = ws.cell(row=row_counter, column=col_num, value=value)  # type: ignore
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )

print("⏳ Starting tender extraction with Claude...")
start_time = time.time()

wb, ws = format_excel(OUTPUT_EXCEL)
row_counter = 2 

processed_count = 0
for filename in os.listdir(TEXT_DIR):
    if not filename.endswith(".json") or processed_count >= MAX_FILES:
        continue

    filepath = os.path.join(TEXT_DIR, filename)
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            tender_json = json.load(f)

        print(f"🔍 Processing ({processed_count+1}/{MAX_FILES}): {filename}")
        text_for_claude = build_tender_text(tender_json)
        result = ask_claude(text_for_claude[:MAX_TOKENS])

        
        try:
            parsed = json.loads(result)
        except json.JSONDecodeError:
            # Try to extract JSON from response
            try:
                start_idx = result.find('{')
                end_idx = result.rfind('}') + 1
                if start_idx != -1 and end_idx != -1:
                    parsed = json.loads(result[start_idx:end_idx])
                else:
                    parsed = {}
            except:
                parsed = {}
        
        # Prepare row data with fallback values
        row_data = {
            "Title": parsed.get("title", "Not extracted"),
            "Issuer": parsed.get("issuer", "Not extracted"),
            "Deadline": parsed.get("deadline", "Not specified"),
            "Budget": parsed.get("budget", "Not specified"),
            "Location": parsed.get("location", "Not specified"),
            "Project Type": parsed.get("project_type", "Not specified"),
            "Required Documents": parsed.get("required_documents", []),
            "PC AVK5 Required": "Yes" if parsed.get("avk5_required") else "No",
            "Technical Specifications": parsed.get("technical_specs", "Not specified"),
            "Payment Terms": parsed.get("payment_terms", "Not specified"),
            "Resource Requirements": parsed.get("resource_requirements", "Not specified"),
            "Timeline Feasibility": parsed.get("timeline_feasibility", "Not assessed"),
            "Profitability Assessment": parsed.get("profitability", "Not assessed"),
            "Filename": filename
        }
        
        save_to_excel(ws, row_data, row_counter)
        row_counter += 1
        processed_count += 1
        
        # API rate limit management
        time.sleep(1.5)
        
    except Exception as e:
        print(f"❌ Error processing {filename}: {str(e)}")
        row_data = {col: "ERROR" for col in COLUMNS}
        row_data["Filename"] = filename
        save_to_excel(ws, row_data, row_counter)
        row_counter += 1

try:
    wb.save(OUTPUT_EXCEL)
    proc_time = time.time() - start_time
    
    print(f"\n✅ Successfully processed {processed_count} tenders")
    print(f"💾 Excel file saved to: {OUTPUT_EXCEL}")
    print(f"⏱️ Total processing time: {proc_time:.2f} seconds")
    print(f"⏳ Average time per tender: {proc_time/processed_count if processed_count else 0:.2f} seconds")
    
except Exception as e:
    print(f"❌ Failed to save Excel file: {str(e)}")

print("🏁 Extraction complete")