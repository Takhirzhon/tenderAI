from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Union, List, Dict, Any

def format_excel(ws):
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    col_widths = [40,30,15,15,20,25,40,15,50,30,40,20,20,30]
    columns = [
        "Назва тендеру","Замовник","Кінцевий термін подачі","Очікуваний бюджет (грн)",
        "Локація проєкту","Тип проєкту","Необхідні документи","Потрібен АВК-5",
        "Технічні вимоги","Умови оплати","Ресурси та матеріали","Реалістичність строків",
        "Оцінка рентабельності","Назва файлу"
    ]

    for col_num, (title, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    return ws

def generate_excel_from_result(
    result: Union[Dict[str, Any], List[Dict[str, Any]]],
    filename: str = "tender.xlsx"
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tender Analysis"
    ws = format_excel(ws)

    # Normalize to list
    rows = result if isinstance(result, list) else [result]

    for row_idx, item in enumerate(rows, start=2):
        values = [
            item.get("title", ""),
            item.get("issuer", ""),
            item.get("deadline", ""),
            item.get("budget", ""),
            item.get("location", ""),
            item.get("project_type", ""),
            ", ".join(item.get("required_documents", [])),
            "✅" if item.get("avk5_required") else "❌",
            item.get("technical_specs", ""),
            item.get("payment_terms", ""),
            item.get("resource_requirements", ""),
            item.get("timeline_feasibility", ""),
            item.get("profitability", ""),
            item.get("filename", "")
        ]
        for col_num, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.alignment = Alignment(wrap_text=True)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
