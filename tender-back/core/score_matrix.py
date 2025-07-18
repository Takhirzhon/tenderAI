# estimation_model.py
import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Ukrainian construction standards database (sample data)
AVK5_STANDARDS = {
    "concrete": {
        "M200": {"unit": "m³", "price": 1800},
        "M300": {"unit": "m³", "price": 2100},
        "M400": {"unit": "m³", "price": 2400}
    },
    "rebar": {
        "A500C-Ø10": {"unit": "ton", "price": 28000},
        "A500C-Ø12": {"unit": "ton", "price": 27500},
        "A500C-Ø16": {"unit": "ton", "price": 27200}
    },
    "labor": {
        "mason": {"unit": "hour", "price": 150},
        "carpenter": {"unit": "hour", "price": 180},
        "electrician": {"unit": "hour", "price": 200}
    },
    "equipment": {
        "crane_25t": {"unit": "shift", "price": 4500},
        "excavator": {"unit": "shift", "price": 3200},
        "concrete_pump": {"unit": "shift", "price": 2800}
    },
    "overhead_rate": 0.15,  # 15% overhead
    "profit_margin": 0.10    # 10% profit
}

class AVK5Estimator:
    """PC AVK5 compliant cost estimation for Ukrainian construction projects"""
    
    def __init__(self, standards=AVK5_STANDARDS):
        self.standards = standards
        
    def calculate_estimate(self, materials, labor, equipment):
        """
        Calculate construction costs according to Ukrainian AVK5 standards
        
        Args:
            materials: dict of {material_type: (quantity, specification)}
            labor: dict of {position: (hours, level)}
            equipment: dict of {equipment_type: (quantity, duration)}
        
        Returns:
            dict: Detailed cost breakdown
        """
        # Material costs
        material_cost = 0
        material_breakdown = []
        for mat_type, (qty, spec) in materials.items():
            if mat_type in self.standards and spec in self.standards[mat_type]:
                unit_price = self.standards[mat_type][spec]["price"]
                cost = qty * unit_price
                material_cost += cost
                material_breakdown.append({
                    "type": mat_type,
                    "specification": spec,
                    "quantity": qty,
                    "unit_price": unit_price,
                    "total": cost
                })
        
        # Labor costs
        labor_cost = 0
        labor_breakdown = []
        for position, (hours, level) in labor.items():
            if position in self.standards["labor"]:
                hourly_rate = self.standards["labor"][position]["price"]
                cost = hours * hourly_rate
                labor_cost += cost
                labor_breakdown.append({
                    "position": position,
                    "hours": hours,
                    "hourly_rate": hourly_rate,
                    "total": cost
                })
        
        # Equipment costs
        equipment_cost = 0
        equipment_breakdown = []
        for equip_type, (qty, duration) in equipment.items():
            if equip_type in self.standards["equipment"]:
                unit_price = self.standards["equipment"][equip_type]["price"]
                cost = qty * duration * unit_price
                equipment_cost += cost
                equipment_breakdown.append({
                    "type": equip_type,
                    "quantity": qty,
                    "duration": duration,
                    "unit_price": unit_price,
                    "total": cost
                })
        
        # Calculate total costs
        direct_costs = material_cost + labor_cost + equipment_cost
        overhead = direct_costs * self.standards["overhead_rate"]
        total_cost = direct_costs + overhead
        profit = total_cost * self.standards["profit_margin"]
        final_price = total_cost + profit
        
        return {
            "material_cost": material_cost,
            "material_breakdown": material_breakdown,
            "labor_cost": labor_cost,
            "labor_breakdown": labor_breakdown,
            "equipment_cost": equipment_cost,
            "equipment_breakdown": equipment_breakdown,
            "direct_costs": direct_costs,
            "overhead": overhead,
            "total_cost": total_cost,
            "profit": profit,
            "final_price": final_price,
            "currency": "UAH"
        }
    
    def export_to_excel(self, estimate, filename):
        """Export cost estimate to Excel with Ukrainian formatting"""
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Кошторис ПК АВК5")
        else:
            ws.title = "Кошторис ПК АВК5"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        
        # Section headers
        sections = [
            ("Матеріали", estimate["material_breakdown"], ["Тип", "Специфікація", "Кількість", "Ціна за од.", "Загалом"]),
            ("Робоча сила", estimate["labor_breakdown"], ["Позиція", "Години", "Годинна ставка", "Загалом"]),
            ("Обладнання", estimate["equipment_breakdown"], ["Тип", "Кількість", "Тривалість", "Ціна за од.", "Загалом"])
        ]
        
        row_idx = 1
        for title, breakdown, headers in sections:
            # Section title
            ws.merge_cells(f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}")
            ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=14)
            row_idx += 1
            
            # Column headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            for item in breakdown:
                row_idx += 1
                for col_idx, value in enumerate(item.values(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add subtotal
            row_idx += 1
            subtotal = {
                "Матеріали": estimate["material_cost"],
                "Робоча сила": estimate["labor_cost"],
                "Обладнання": estimate["equipment_cost"]
            }[title]
            
            ws.cell(row=row_idx, column=len(headers)-1, value="Підсумок:").font = Font(bold=True)
            ws.cell(row=row_idx, column=len(headers), value=subtotal).font = Font(bold=True)
            row_idx += 2
        
        # Summary section
        summary_data = [
            ("Прямі витрати", estimate["direct_costs"]),
            ("Накладні витрати (15%)", estimate["overhead"]),
            ("Загальна собівартість", estimate["total_cost"]),
            ("Прибуток (10%)", estimate["profit"]),
            ("ФІНАЛЬНА ЦІНА", estimate["final_price"])
        ]
        
        ws.cell(row=row_idx, column=1, value="ЗВЕДЕНИЙ КОШТОРИС").font = Font(bold=True, size=14)
        row_idx += 1
        
        for label, value in summary_data:
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value).number_format = '#,##0.00'
            if "ФІНАЛЬНА" in label:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FF0000")
                ws.cell(row=row_idx, column=2).font = Font(bold=True, color="FF0000")
            row_idx += 1
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width # type: ignore   
        
        wb.save(filename)
        return filename

class DocumentComplianceChecker:
    """Check tender document requirements against company's document vault"""
    
    def __init__(self, document_vault_path="../data/document_vault.json"):
        self.vault_path = document_vault_path
        self.document_vault = self.load_document_vault()
    
    def load_document_vault(self):
        """Load company document metadata from JSON file"""
        if os.path.exists(self.vault_path):
            with open(self.vault_path, "r", encoding="utf-8") as f:
                return json.load(f)
        return {
            "documents": [],
            "categories": {},
            "last_updated": datetime.now().isoformat()
        }
    
    def save_document_vault(self):
        os.makedirs(os.path.dirname(self.vault_path), exist_ok=True)
        """Save document vault to file"""
        with open(self.vault_path, "w", encoding="utf-8") as f:
            json.dump(self.document_vault, f, ensure_ascii=False, indent=2)
    
    def add_document(self, doc_name, doc_type, validity, file_path, tags=None):
        """Add document to company vault"""
        new_doc = {
            "id": f"DOC-{len(self.document_vault['documents']) + 1:04d}",
            "name": doc_name,
            "type": doc_type,
            "validity": validity,
            "path": file_path,
            "tags": tags or [],
            "added_date": datetime.now().isoformat()
        }
        self.document_vault['documents'].append(new_doc)
        
        # Update categories
        if doc_type not in self.document_vault['categories']:
            self.document_vault['categories'][doc_type] = []
        self.document_vault['categories'][doc_type].append(new_doc['id'])
        
        self.save_document_vault()
        return new_doc
    
    def check_compliance(self, required_docs):
        """
        Check compliance with tender document requirements
        
        Args:
            required_docs: list of required document types
        
        Returns:
            dict: compliance report with missing documents
        """
        available_types = set(doc['type'] for doc in self.document_vault['documents'])
        
        missing = []
        for doc_type in required_docs:
            if doc_type not in available_types:
                missing.append(doc_type)
        
        compliance_score = 1 - (len(missing) / len(required_docs)) if required_docs else 1.0
        
        return {
            "required_documents": required_docs,
            "available_documents": list(available_types),  # Convert set to list
            "missing_documents": missing,
            "compliance_score": compliance_score,
            "is_compliant": len(missing) == 0
        }
    
    def suggest_alternatives(self, missing_docs):
        """Suggest similar documents for missing requirements"""
        suggestions = {}
        doc_types = set(doc['type'] for doc in self.document_vault['documents'])
        
        for doc in missing_docs:
            # Simple matching - in real app use NLP similarity
            matches = [t for t in doc_types if doc.lower() in t.lower()]
            suggestions[doc] = matches
        
        return suggestions

class ProfitabilityAnalyzer:
    """Analyze tender profitability considering costs, risks, and timeline"""
    
    def __init__(self, avk5_estimator):
        self.estimator = avk5_estimator
    
    def analyze_tender(self, tender_data, company_resources):
        """
        Comprehensive profitability analysis for a tender

        Args:
            tender_data: dict with tender details (budget, resources, etc.)
            company_resources: dict with company capabilities

        Returns:
            dict: enriched profitability analysis report
        """
        # Estimate cost from materials, labor, and equipment
        cost_estimate = self.estimator.calculate_estimate(
            tender_data.get("materials", {}),
            tender_data.get("labor", {}),
            tender_data.get("equipment", {})
        )
        estimated_cost = cost_estimate["final_price"]

        # Extract tender value safely
        tender_value = tender_data.get("budget", 0)
        try:
            tender_value = float(str(tender_value).replace(",", "").split()[0])
        except:
            tender_value = 0.0

        gross_profit = tender_value - estimated_cost
        profit_margin = gross_profit / tender_value if tender_value else 0

        # Resource gap
        resource_gap = self.analyze_resource_gap(
            tender_data.get("resource_requirements", {}),
            company_resources
        )

        # Timeline feasibility
        timeline_feasibility = self.assess_timeline(
            tender_data.get("timeline", {}),
            company_resources.get("current_projects", [])
        )

        # Risk scoring
        risk_factors = self.assess_risks(tender_data)

        # ROI Score
        roi_score = self.calculate_roi_score(
            profit_margin,
            risk_factors.get("composite_risk", 0),
            resource_gap.get("resource_availability_score", 0),
            timeline_feasibility.get("feasibility_score", 0)
        )

        return {
            "tender_value": tender_value,
            "estimated_cost": estimated_cost,
            "gross_profit": gross_profit,
            "profit_margin": profit_margin,
            "cost_breakdown": cost_estimate,
            "resource_gap": resource_gap,
            "timeline_feasibility": timeline_feasibility,
            "risk_factors": risk_factors,
            "roi_score": roi_score,
            "recommendation": "BID" if roi_score >= 70 else "NO-BID"
        }

    
    def analyze_resource_gap(self, requirements, resources):
        """Analyze gap between required and available resources"""
        gap_analysis = {}
        availability_score = 0
        total_resources = 0
        
        for resource_type, required in requirements.items():
            available = resources.get(resource_type, 0)
            gap = required - available
            gap_percent = gap / required if required else 0
            
            gap_analysis[resource_type] = {
                "required": required,
                "available": available,
                "gap": gap,
                "gap_percent": gap_percent
            }
            
            if gap <= 0:
                availability_score += 1
            total_resources += 1
        
        return {
            "gap_analysis": gap_analysis,
            "resource_availability_score": (availability_score / total_resources) * 100 if total_resources else 100
        }
    
    def assess_timeline(self, timeline, current_projects):
        """Assess timeline feasibility considering current workload"""
        required_duration = timeline.get("duration_days", 0)
        start_date = timeline.get("start_date")
        
        # Simplified - in real app use project scheduling
        current_workload = sum(proj["duration"] for proj in current_projects)
        feasibility_score = max(0, 100 - (current_workload / 30) * 100)  # Assume 30 days max capacity
        
        return {
            "required_duration": required_duration,
            "current_workload": current_workload,
            "feasibility_score": feasibility_score,
            "is_feasible": feasibility_score >= 70
        }
    
    def assess_risks(self, tender_data):
        """Assess project risks based on tender details"""
        risks = {
            "technical_complexity": tender_data.get("complexity", 0) * 0.3,
            "payment_terms": 0.4 if tender_data.get("payment_terms") == "deferred" else 0.1,
            "penalty_clauses": 0.2 if tender_data.get("has_penalties") else 0.05,
            "competition_level": min(1.0, tender_data.get("competitors", 0) * 0.1)
        }
        
        composite_risk = sum(risks.values()) / len(risks) if risks else 0
        risks["composite_risk"] = composite_risk
        
        return risks
    
    def calculate_roi_score(self, profit_margin, risk, availability, timeline):
        """Calculate composite ROI score (0-100)"""
        # Weighted factors
        return (
            (profit_margin * 100 * 0.5) + 
            (availability * 0.2) + 
            (timeline * 0.2) + 
            ((1 - risk) * 100 * 0.1)
        )

# Example Usage
if __name__ == "__main__":
    # Initialize modules
    avk5 = AVK5Estimator()
    compliance = DocumentComplianceChecker()
    profitability = ProfitabilityAnalyzer(avk5)
    
    # Example tender data
    tender = {
        "title": "School Renovation Project",
        "budget": 2500000,
        "materials": {
            "concrete": (120, "M300"),
            "rebar": (8, "A500C-Ø12")
        },
        "labor": {
            "mason": (240, "standard"),
            "carpenter": (120, "standard")
        },
        "equipment": {
            "crane_25t": (1, 5),
            "concrete_pump": (1, 3)
        },
        "resource_requirements": {
            "workers": 15,
            "engineers": 2,
            "vehicles": 3
        },
        "timeline": {
            "duration_days": 90,
            "start_date": "2025-09-01"
        },
        "complexity": 7,  # 1-10 scale
        "payment_terms": "deferred",
        "has_penalties": True,
        "competitors": 5,
        "required_docs": ["License", "Tax Certificate", "Experience Portfolio"]
    }
    
    # Company resources
    company = {
        "workers": 12,
        "engineers": 3,
        "vehicles": 2,
        "current_projects": [
            {"name": "Hospital Project", "duration": 45},
            {"name": "Apartment Building", "duration": 30}
        ]
    }
    
    # Add documents to vault
    compliance.add_document("Construction License", "License", "2026-12-31", "/docs/license.pdf")
    compliance.add_document("Tax Certificate 2025", "Tax Certificate", "2025-12-31", "/docs/tax_cert.pdf")
    
    # Run analyses
    cost_estimate = avk5.calculate_estimate(
        tender["materials"], tender["labor"], tender["equipment"]
    )
    doc_compliance = compliance.check_compliance(tender["required_docs"])
    profit_analysis = profitability.analyze_tender(tender, company)
    
    # Export results
    avk5.export_to_excel(cost_estimate, "cost_estimate.xlsx")
    
    print("Cost Estimate:")
    print(json.dumps(cost_estimate, indent=2))
    
    print("\nDocument Compliance:")
    print(json.dumps(doc_compliance, indent=2))
    
    print("\nProfitability Analysis:")
    print(json.dumps(profit_analysis, indent=2))