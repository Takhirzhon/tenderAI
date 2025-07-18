import os
from openpyxl import Workbook
from core.extract_to_excel import format_excel
from core.data_extractor import extract_text_from_pdf
from core.analyze_tender import analyze_tender
import streamlit as st

UPLOAD_DIR = "../uploaded/"
TEXT_DIR = "../extracted/"

def process_all_pdfs(client):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Аналіз тендерів")
    else:
        ws.title = "Аналіз тендерів"
    ws, columns = format_excel(ws)

    row = 2
    for filename in os.listdir(UPLOAD_DIR):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(UPLOAD_DIR, filename)
            try:
                text = extract_text_from_pdf(pdf_path)
                result = analyze_tender(text, client)

                if result:
                    for col_index, key in enumerate([
                        "title", "issuer", "deadline", "budget", "location",
                        "project_type", "required_documents", "avk5_required",
                        "technical_specs", "payment_terms", "resource_requirements",
                        "timeline_feasibility", "profitability"
                    ], start=1):
                        value = result.get(key, "")
                        if isinstance(value, list):
                            value = ", ".join(value)
                        elif isinstance(value, bool):
                            value = str(value)
                        ws.cell(row=row, column=col_index, value=value)
                    ws.cell(row=row, column=len(columns), value=filename)
                    row += 1
            except Exception as e:
                st.warning(f"⚠️ Error processing {filename}: {e}")

    output_path = os.path.join(TEXT_DIR, "tender_analysis.xlsx")
    wb.save(output_path)
    return output_path