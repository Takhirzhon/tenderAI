import re
import streamlit as st
from datetime import datetime, timedelta
import json
import os
import sys
import time
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from core.score_matrix import AVK5Estimator, DocumentComplianceChecker, ProfitabilityAnalyzer
import xml.etree.ElementTree as ET
from core.uploader import handle_uploaded_tender
from core.downloader import download_prozorro_tenders
from core.data_extractor import extract_text_from_pdf
from core.analyze_link import analyze_tender_from_link

UPLOAD_DIR = "../uploaded/"
TEXT_DIR = "../extracted/"

# Initialize Claude client
def get_claude_client():
    api_key = os.getenv("CLAUDE_API_KEY")
    if not api_key:
        st.error("‚ùå –ö–ª—é—á API Claude –Ω–µ –∑–Ω–∞–π–¥–µ–Ω–æ —É —Ñ–∞–π–ª—ñ .env")
        return None
    return anthropic.Anthropic(api_key=api_key)

# Enhanced parser function
def analyze_tender(text, client):
    prompt = f"""
–í–∏ ‚Äî –µ–∫—Å–ø–µ—Ä—Ç –∑ –¥–µ—Ä–∂–∞–≤–Ω–∏—Ö –∑–∞–∫—É–ø—ñ–≤–µ–ª—å –≤ –£–∫—Ä–∞—ó–Ω—ñ. –ü—Ä–æ–∞–Ω–∞–ª—ñ–∑—É–π—Ç–µ –Ω–∞–≤–µ–¥–µ–Ω–∏–π –Ω–∏–∂—á–µ —Ç–µ–∫—Å—Ç —Ç–µ–Ω–¥–µ—Ä—É —Ç–∞ –≤–∏—Ç—è–≥–Ω—ñ—Ç—å –Ω–µ–æ–±—Ö—ñ–¥–Ω—ñ –ø–æ–ª—è. 
–ü–æ–≤–µ—Ä–Ω—ñ—Ç—å **—Ç—ñ–ª—å–∫–∏ –≤–∞–ª—ñ–¥–Ω–∏–π JSON-–æ–±'—î–∫—Ç** –∑ —Ç–∞–∫–∏–º–∏ –∫–ª—é—á–∞–º–∏:

- title (string): –ü–æ–≤–Ω–∞ –Ω–∞–∑–≤–∞ —Ç–µ–Ω–¥–µ—Ä—É.
- issuer (string): –ù–∞–∑–≤–∞ –∑–∞–º–æ–≤–Ω–∏–∫–∞ –∞–±–æ –æ—Ä–≥–∞–Ω—ñ–∑–∞—Ç–æ—Ä–∞ –∑–∞–∫—É–ø—ñ–≤–ª—ñ.
- deadline (string): –î–µ–¥–ª–∞–π–Ω –ø–æ–¥–∞—á—ñ –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ–π –∞–±–æ –∫—ñ–Ω—Ü–µ–≤–∞ –¥–∞—Ç–∞.
- budget (string): –û—á—ñ–∫—É–≤–∞–Ω–∞ –≤–∞—Ä—Ç—ñ—Å—Ç—å –∑–∞–∫—É–ø—ñ–≤–ª—ñ –∞–±–æ –≥—Ä–∞–Ω–∏—á–Ω–∞ —Å—É–º–∞ –∫–æ–Ω—Ç—Ä–∞–∫—Ç—É.
- location (string): –ú—ñ—Å—Ü–µ —Ä–µ–∞–ª—ñ–∑–∞—Ü—ñ—ó –∞–±–æ –≤–∏–∫–æ–Ω–∞–Ω–Ω—è —Ä–æ–±—ñ—Ç.
- project_type (string): –¢–∏–ø –ø—Ä–æ—î–∫—Ç—É (–Ω–∞–ø—Ä–∏–∫–ª–∞–¥: –±—É–¥—ñ–≤–Ω–∏—Ü—Ç–≤–æ, IT-–ø–æ—Å–ª—É–≥–∏, —Ä–µ–º–æ–Ω—Ç).
- required_documents (list of strings): –£–∫–∞–∂—ñ—Ç—å **—É—Å—ñ –¥–æ–∫—É–º–µ–Ω—Ç–∏**, —è–∫—ñ –≤–∏–º–∞–≥–∞—é—Ç—å—Å—è –¥–ª—è —É—á–∞—Å—Ç—ñ –∞–±–æ –∫–≤–∞–ª—ñ—Ñ—ñ–∫–∞—Ü—ñ—ó. –ü—Ä–∏–∫–ª–∞–¥–∏: –í–∏—Ç—è–≥ –∑ –Ñ–î–†, –≥–∞—Ä–∞–Ω—Ç—ñ—è —Ç–µ–Ω–¥–µ—Ä–Ω–æ—ó –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ—ó, AVK-5 –∫–æ—à—Ç–æ—Ä–∏—Å, –¥–æ–≤—ñ–¥–∫–∞ –ø—Ä–æ –¥–æ—Å–≤—ñ–¥, –¥–æ–∫—É–º–µ–Ω—Ç–∏ –Ω–∞ —Ç–µ—Ö–Ω—ñ–∫—É, –¥–µ–∫–ª–∞—Ä–∞—Ü—ñ—è –ø—Ä–æ –ø–µ—Ä—Å–æ–Ω–∞–ª. –û—Ä—ñ—î–Ω—Ç—É–π—Ç–µ—Å—å –Ω–∞ —Ä–æ–∑–¥—ñ–ª–∏ –∑ –Ω–∞–∑–≤–∞–º–∏: ‚Äú–ö–≤–∞–ª—ñ—Ñ—ñ–∫–∞—Ü—ñ–π–Ω—ñ –≤–∏–º–æ–≥–∏‚Äù, ‚Äú–ù–µ–æ–±—Ö—ñ–¥–Ω—ñ –¥–æ–∫—É–º–µ–Ω—Ç–∏‚Äù, ‚Äú–ü–µ—Ä–µ–ª—ñ–∫ –¥–æ–∫—É–º–µ–Ω—Ç—ñ–≤‚Äù.
- avk5_required (boolean): –ß–∏ –ø—Ä—è–º–æ –∑–∞–∑–Ω–∞—á–µ–Ω–∞ –≤–∏–º–æ–≥–∞ –ø–æ–¥–∞—Ç–∏ –∫–æ—à—Ç–æ—Ä–∏—Å —É —Ñ–æ—Ä–º–∞—Ç—ñ –ê–í–ö-5 (true –∞–±–æ false).
- technical_specs (string): –°—Ç–∏—Å–ª–µ —Ä–µ–∑—é–º–µ —Ç–µ—Ö–Ω—ñ—á–Ω–∏—Ö –≤–∏–º–æ–≥ ‚Äî —â–æ –ø–æ—Ç—Ä—ñ–±–Ω–æ –≤–∏–∫–æ–Ω–∞—Ç–∏, —è–∫—ñ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏, —Å—Ç–∞–Ω–¥–∞—Ä—Ç–∏, –æ–±–ª–∞–¥–Ω–∞–Ω–Ω—è. –û—Ä—ñ—î–Ω—Ç—É–π—Ç–µ—Å—å –Ω–∞ —Ä–æ–∑–¥—ñ–ª–∏ ‚Äú–¢–µ—Ö–Ω—ñ—á–Ω–∞ —Å–ø–µ—Ü–∏—Ñ—ñ–∫–∞—Ü—ñ—è‚Äù, ‚Äú–¢–µ—Ö–Ω—ñ—á–Ω–µ –∑–∞–≤–¥–∞–Ω–Ω—è‚Äù, ‚Äú–û–ø–∏—Å –ø—Ä–µ–¥–º–µ—Ç—É –∑–∞–∫—É–ø—ñ–≤–ª—ñ‚Äù.
- payment_terms (string): –£–º–æ–≤–∏ –æ–ø–ª–∞—Ç–∏ ‚Äî –∞–≤–∞–Ω—Å, –ø–æ–µ—Ç–∞–ø–Ω–∞, –ø—ñ—Å–ª—è –≤–∏–∫–æ–Ω–∞–Ω–Ω—è, —Å—Ç—Ä–æ–∫–∏ —Ç–∞ –ø–æ—Ä—è–¥–æ–∫.
- resource_requirements (string): –£–∫–∞–∂—ñ—Ç—å, —è–∫—ñ **—Ä–µ—Å—É—Ä—Å–∏ —Ç–∞ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏** –ø–æ–≤–∏–Ω–µ–Ω –Ω–∞–¥–∞—Ç–∏ –ø—ñ–¥—Ä—è–¥–Ω–∏–∫: —Ä–æ–±—ñ—Ç–Ω–∏–∫–∏, —ñ–Ω–∂–µ–Ω–µ—Ä–∏, —Ç–µ—Ö–Ω—ñ–∫–∞, –±—É–¥–º–∞—Ç–µ—Ä—ñ–∞–ª–∏, –ü–ú–ú —Ç–æ—â–æ. –û—Ä—ñ—î–Ω—Ç—É–π—Ç–µ—Å—å –Ω–∞ —Ä–æ–∑–¥—ñ–ª–∏ ‚Äú–û–±—Å—è–≥ —Ä–æ–±—ñ—Ç‚Äù, ‚Äú–í–∏–º–æ–≥–∏ –¥–æ —Ä–µ—Å—É—Ä—Å—ñ–≤‚Äù, ‚Äú–í—ñ–¥–æ–º—ñ—Å—Ç—å —Ä–µ—Å—É—Ä—Å—ñ–≤‚Äù.
- timeline_feasibility (string): –ß–∏ –≤–∏–≥–ª—è–¥–∞—é—Ç—å —Å—Ç—Ä–æ–∫–∏ –≤–∏–∫–æ–Ω–∞–Ω–Ω—è —Ä–µ–∞–ª—ñ—Å—Ç–∏—á–Ω–∏–º–∏ (–Ω–∞–ø—Ä–∏–∫–ª–∞–¥: ‚Äú—Ä–µ–∞–ª—ñ—Å—Ç–∏—á–Ω—ñ‚Äù, ‚Äú–æ–±–º–µ–∂–µ–Ω—ñ, –∞–ª–µ –º–æ–∂–ª–∏–≤—ñ‚Äù, ‚Äú–Ω–µ—Ä–µ–∞–ª—ñ—Å—Ç–∏—á–Ω—ñ‚Äù).
- profitability (string): –ß–∏ –≤–∏–≥–ª—è–¥–∞—î —É—á–∞—Å—Ç—å —É —Ç–µ–Ω–¥–µ—Ä—ñ –ø—Ä–∏–±—É—Ç–∫–æ–≤–æ—é –∑ —É—Ä–∞—Ö—É–≤–∞–Ω–Ω—è–º –±—é–¥–∂–µ—Ç—É, –æ–±—Å—è–≥—É —Ä–æ–±—ñ—Ç —Ç–∞ –≤–∏—Ç—Ä–∞—Ç. –ü—Ä–∏–∫–ª–∞–¥–∏: ‚Äú–ø—Ä–∏–±—É—Ç–∫–æ–≤–∏–π‚Äù, ‚Äú—Ä–∏–∑–∏–∫–æ–≤–∞–Ω–∏–π‚Äù, ‚Äú–∑–±–∏—Ç–∫–æ–≤–∏–π‚Äù ‚Äî –∑ –∫–æ—Ä–æ—Ç–∫–∏–º –ø–æ—è—Å–Ω–µ–Ω–Ω—è–º.

‚ùó–ù–µ –¥–æ–¥–∞–≤–∞–π—Ç–µ –ø–æ—è—Å–Ω–µ–Ω—å –∞–±–æ –∫–æ–º–µ–Ω—Ç–∞—Ä—ñ–≤ ‚Äî –ø–æ–≤–µ—Ä–Ω—ñ—Ç—å –ª–∏—à–µ –∫–æ—Ä–µ–∫—Ç–Ω–∏–π JSON.

Tender Text:
\"\"\"{text[:15000]}\"\"\"
"""


    try:
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1024,
            temperature=0.0,
            system = "–í–∏ ‚Äî –∞–Ω–∞–ª—ñ—Ç–∏–∫ –¥–µ—Ä–∂–∞–≤–Ω–∏—Ö –∑–∞–∫—É–ø—ñ–≤–µ–ª—å. –í–∞—à–µ –∑–∞–≤–¥–∞–Ω–Ω—è ‚Äî —Å—Ç—Ä—É–∫—Ç—É—Ä—É–≤–∞—Ç–∏ –∫–ª—é—á–æ–≤—ñ –ø–∞—Ä–∞–º–µ—Ç—Ä–∏ —Ç–µ–Ω–¥–µ—Ä—É: –¥–æ–∫—É–º–µ–Ω—Ç–∏, —Ç–µ—Ö–Ω—ñ—á–Ω—ñ –≤–∏–º–æ–≥–∏, —Ä–µ—Å—É—Ä—Å–∏, —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω—ñ—Å—Ç—å, –ê–í–ö-5. –í–∏—Ö—ñ–¥ –º–∞—î –±—É—Ç–∏ —É –≤–∏–≥–ª—è–¥—ñ –≤–∞–ª—ñ–¥–Ω–æ–≥–æ JSON.",
            messages=[{"role": "user", "content": prompt}]
        )
        if response.content:
            result = ''.join(block.text for block in response.content if hasattr(block, 'text'))
            try:
                return json.loads(result)
            except json.JSONDecodeError:
                match = re.search(r'({.*})', result, re.DOTALL) # type: ignore
                if match:
                    return json.loads(match.group(1))
        st.error("‚ùå –ö–ª–æ–¥ –ø–æ–≤–µ—Ä–Ω—É–≤ –Ω–µ–¥—ñ–π—Å–Ω–∏–π JSON.")
        return {}
    except Exception as e:
        st.error(f"‚ùå –ü–æ–º–∏–ª–∫–∞ –∞–Ω–∞–ª—ñ–∑—É —Ç–µ–Ω–¥–µ—Ä—É: {e}")
        return {}

def format_excel(ws):
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    col_widths = [40, 30, 15, 15, 20, 25, 40, 15, 50, 30, 40, 20, 20, 30]
    columns = [
    "–ù–∞–∑–≤–∞ —Ç–µ–Ω–¥–µ—Ä—É",
    "–ó–∞–º–æ–≤–Ω–∏–∫",
    "–ö—ñ–Ω—Ü–µ–≤–∏–π —Ç–µ—Ä–º—ñ–Ω –ø–æ–¥–∞—á—ñ",
    "–û—á—ñ–∫—É–≤–∞–Ω–∏–π –±—é–¥–∂–µ—Ç (–≥—Ä–Ω)",
    "–õ–æ–∫–∞—Ü—ñ—è –ø—Ä–æ—î–∫—Ç—É",
    "–¢–∏–ø –ø—Ä–æ—î–∫—Ç—É",
    "–ù–µ–æ–±—Ö—ñ–¥–Ω—ñ –¥–æ–∫—É–º–µ–Ω—Ç–∏",
    "–ü–æ—Ç—Ä—ñ–±–µ–Ω –ê–í–ö-5",
    "–¢–µ—Ö–Ω—ñ—á–Ω—ñ –≤–∏–º–æ–≥–∏",
    "–£–º–æ–≤–∏ –æ–ø–ª–∞—Ç–∏",
    "–†–µ—Å—É—Ä—Å–∏ —Ç–∞ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏",
    "–†–µ–∞–ª—ñ—Å—Ç–∏—á–Ω—ñ—Å—Ç—å —Å—Ç—Ä–æ–∫—ñ–≤",
    "–û—Ü—ñ–Ω–∫–∞ —Ä–µ–Ω—Ç–∞–±–µ–ª—å–Ω–æ—Å—Ç—ñ",
    "–ù–∞–∑–≤–∞ —Ñ–∞–π–ª—É"
]


    
    for col_num, (column_title, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width
    return ws, columns

def process_all_pdfs(client):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("–ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤")
    else:
        ws.title = "–ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤"
    ws, columns = format_excel(ws)

    row = 2
    for filename in os.listdir(UPLOAD_DIR):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(UPLOAD_DIR, filename)
            try:
                text = extract_text_from_pdf(pdf_path)
                result = analyze_tender(text, client)

                if result:
                    for col_index, key in enumerate([
                        "title", "issuer", "deadline", "budget", "location",
                        "project_type", "required_documents", "avk5_required",
                        "technical_specs", "payment_terms", "resource_requirements",
                        "timeline_feasibility", "profitability"
                    ], start=1):
                        value = result.get(key, "")
                        if isinstance(value, list):
                            value = ", ".join(value)
                        elif isinstance(value, bool):
                            value = str(value)
                        ws.cell(row=row, column=col_index, value=value)
                    ws.cell(row=row, column=len(columns), value=filename)
                    row += 1
            except Exception as e:
                st.warning(f"‚ö†Ô∏è Error processing {filename}: {e}")

    output_path = os.path.join(TEXT_DIR, "tender_analysis.xlsx")
    wb.save(output_path)
    return output_path

# Streamlit App
st.set_page_config(page_title="AI –û–ø—Ç–∏–º—ñ–∑–∞—Ç–æ—Ä —Ç–µ–Ω–¥–µ—Ä—ñ–≤", layout="wide")

# Initialize session state
if "analysis_results" not in st.session_state:
    st.session_state.analysis_results = []
if "tenders_downloaded" not in st.session_state:
    st.session_state.tenders_downloaded = []
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
if "analysis_attempted" not in st.session_state:
    st.session_state.analysis_attempted = False
if "company_resources" not in st.session_state:
    st.session_state.company_resources = {
        "workers": 10,
        "engineers": 2,
        "vehicles": 3,
        "current_projects": []
    }
if "document_vault" not in st.session_state:
    st.session_state.document_vault = DocumentComplianceChecker()

# Sidebar
st.sidebar.title("üõ†Ô∏è –ù–∞–ª–∞—à—Ç—É–≤–∞–Ω–Ω—è —Ç–µ–Ω–¥–µ—Ä–∞")
tab = st.sidebar.radio("Navigation", ["üì• –ó–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–Ω—è –¥–∞–Ω–∏—Ö", "üîç –ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤", "üè¢ –ü—Ä–æ—Ñ—ñ–ª—å –∫–æ–º–ø–∞–Ω—ñ—ó", "üìä –û—Ü—ñ–Ω–∫–∞ —Ç–µ–Ω–¥–µ—Ä–Ω–∏—Ö –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ–π"])

keywords_option = st.sidebar.multiselect(
    "üîç Select Tender Topics",
    ["IT", "Construction", "Medical Equipment", "Transport", "Education"],
    default=["Construction"]
)

days_range = st.sidebar.slider(
    "üìÖ Days Back to Search",
    min_value=1,
    max_value=60,
    value=30
)

# Main Page
st.title("üì¶ AI –û–ø—Ç–∏–º—ñ–∑–∞—Ç–æ—Ä —Ç–µ–Ω–¥–µ—Ä—ñ–≤")

if tab == "üì• –ó–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–Ω—è –¥–∞–Ω–∏—Ö":

    handle_uploaded_tender()
    link = st.text_input("üîó –í–≤–µ–¥—ñ—Ç—å –ø–æ—Å–∏–ª–∞–Ω–Ω—è –Ω–∞ —Ç–µ–Ω–¥–µ—Ä")

    if link:
        try:
            result = analyze_tender_from_link(link)
            st.success("‚úÖ –ê–Ω–∞–ª—ñ–∑ –∑–∞–≤–µ—Ä—à–µ–Ω–æ!")
            st.json(result)
        except ValueError as ve:
            st.error(str(ve))
        except RuntimeError as re:
            st.error(str(re))


    st.header("üì• –ó–∞–≤–∞–Ω—Ç–∞–∂—Ç–µ —Ç–µ–Ω–¥–µ—Ä–Ω—ñ –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ—ó –∑ ProZorro ")
    st.write("–Ø–∫—â–æ –≤–∏ —Ö–æ—á–µ—Ç–µ –∞–Ω–∞–ª—ñ–∑—É–≤–∞—Ç–∏ –≤–µ–ª–∏–∫—É –∫—ñ–ª—å–∫—ñ—Å—Ç—å –≤–∏–ø–∞–¥–∫–æ–≤–∏—Ö —Ç–µ–Ω–¥–µ—Ä—ñ–≤")
    # Load topics from keywords.json dynamically
    with open("C:/Users/tashmatov/tender/data/keywords.json", "r", encoding="utf-8") as f:
        topic_keywords = json.load(f)
    topic_list = list(topic_keywords.keys())

    topic = st.selectbox("üìö –í–∏–±–µ—Ä—ñ—Ç—å —Ç–µ–º—É —Ç–µ–Ω–¥–µ—Ä–∞", topic_list)
    num_tenders = st.slider("üì¶ –ö—ñ–ª—å–∫—ñ—Å—Ç—å —Ç–µ–Ω–¥–µ—Ä—ñ–≤ –¥–ª—è –∑–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–Ω—è", min_value=5, max_value=100, value=20, step=5)
    days_back = st.slider("üìÖ –ü–æ—à—É–∫ —Ç–µ–Ω–¥–µ—Ä—ñ–≤ –∑–∞ –æ—Å—Ç–∞–Ω–Ω—ñ N –¥–Ω—ñ–≤", min_value=1, max_value=60, value=30)

    if st.button("üöÄ –ü–æ—á–∞—Ç–∏ –∑–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–Ω—è"):
        with st.spinner("–ó–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–Ω—è..."):
            tenders = download_prozorro_tenders(
                topic=topic,
                total_to_download=num_tenders,
                days_back=days_back
            )

        if tenders:
            st.success(f"‚úÖ {len(tenders)} —Ç–µ–Ω–¥–µ—Ä—ñ–≤ –∑–∞–≤–∞–Ω—Ç–∞–∂–µ–Ω–æ –¥–ª—è —Ç–µ–º–∏ '{topic}'")
            st.session_state["tenders_downloaded"] = tenders

            with st.expander("üìÑ –°—É–º–∞—Ä–Ω–∞ —ñ–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—è –ø—Ä–æ —Ç–µ–Ω–¥–µ—Ä–∏"):
                st.json({
                    "topic": topic,
                    "keywords": topic_keywords[topic],
                    "total_downloaded": len(tenders)
                })

            st.download_button(
                label="üìÅ –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –º–µ—Ç–∞–¥–∞–Ω—ñ —Ç–µ–Ω–¥–µ—Ä—ñ–≤",
                data=json.dumps(tenders, ensure_ascii=False, indent=2),
                file_name=f"{topic.lower()}_tenders_summary.json",
                mime="application/json"
            )
        else:
            st.warning("‚ùå –ù–µ –∑–Ω–∞–π–¥–µ–Ω–æ —Ç–µ–Ω–¥–µ—Ä—ñ–≤, —â–æ –≤—ñ–¥–ø–æ–≤—ñ–¥–∞—é—Ç—å –≤–∞—à–∏–º –∫—Ä–∏—Ç–µ—Ä—ñ—è–º.")

elif tab == "üîç –ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤":
    st.header("üîç –ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤")
    # Check for existing tender files
    def load_existing_tenders():
        tender_files = []
        tenders_dir = "../tenders"
        uploaded_dir = "../uploaded"
        if os.path.exists(tenders_dir):
            for filename in os.listdir(tenders_dir):
                if filename.startswith("ProZorro_") and filename.endswith(".json"):
                    tender_id = filename.replace("ProZorro_", "").replace(".json", "")
                    try:
                        with open(os.path.join(tenders_dir, filename), "r", encoding="utf-8") as f:
                            tender_data = json.load(f)
                            tender_files.append({
                                "id": tender_id,
                                "title": tender_data.get("title", "–ë–µ–∑ –Ω–∞–∑–≤–∏"),
                                "date": tender_data.get("dateModified", ""),
                                "budget": tender_data.get("value", {}).get("amount", 0),
                                "file": filename
                            })
                    except Exception as e:
                        st.warning(f"Error loading {filename}: {e}")
        if os.path.exists(uploaded_dir):
            for filename in os.listdir(uploaded_dir):
                if filename.endswith(".pdf"):
                    tender_files.append({
                        "id": filename.replace(".pdf", ""),
                        "title": filename.replace(".pdf", ""),
                        "date": "",
                        "budget": 0,
                        "file": filename
                    })
        return tender_files

    # Load tenders if missing
    if not st.session_state.tenders_downloaded:
        existing = load_existing_tenders()
        
        # Also include analyzed uploaded tenders from session state
        uploaded = []
        for result in st.session_state.analysis_results:
            if "tender_id" in result and "title" in result:
                uploaded.append({
                    "id": result["tender_id"],
                    "title": result["title"],
                    "date": result.get("deadline", ""),
                    "budget": result.get("budget", 0),
                    "file": result.get("Filename", "uploaded.pdf")
                })

        combined = existing + uploaded
        if combined:
            st.session_state.tenders_downloaded = combined
            st.success(f"‚úÖ Loaded {len(combined)} tenders (existing + uploaded).")
        else:
            st.warning("‚ö†Ô∏è –ù–µ –∑–Ω–∞–π–¥–µ–Ω–æ —Ç–µ–Ω–¥–µ—Ä–Ω–∏—Ö —Ñ–∞–π–ª—ñ–≤.")
            st.stop()


    tender_options = {t['id']: t['title'] for t in st.session_state.tenders_downloaded}
    selected_tenders = st.multiselect(
        "–í–∏–±–µ—Ä—ñ—Ç—å —Ç–µ–Ω–¥–µ—Ä–∏ –¥–ª—è –∞–Ω–∞–ª—ñ–∑—É:",
        options=list(tender_options.keys()),
        format_func=lambda x: f"{x} - {tender_options[x][:50]}..."
    )

    if not selected_tenders:
        st.info("‚ÑπÔ∏è –ë—É–¥—å –ª–∞—Å–∫–∞, –≤–∏–±–µ—Ä—ñ—Ç—å –ø—Ä–∏–Ω–∞–π–º–Ω—ñ –æ–¥–∏–Ω —Ç–µ–Ω–¥–µ—Ä –¥–ª—è –∞–Ω–∞–ª—ñ–∑—É.")
        st.stop()

    client = get_claude_client()
    if not client:
        st.stop()

    status_text = st.empty()
    analyze_clicked = st.button("üîç –ê–Ω–∞–ª—ñ–∑ –≤–∏–±—Ä–∞–Ω–∏—Ö —Ç–µ–Ω–¥–µ—Ä—ñ–≤", key="analyze_button")

    if analyze_clicked:
        st.session_state.analysis_attempted = True
        results = []
        progress_bar = st.progress(0)

        for i, tid in enumerate(selected_tenders):
            tender = next((t for t in st.session_state.tenders_downloaded if t['id'] == tid), None)
            if not tender:
                continue

            progress_bar.progress((i + 1) / len(selected_tenders))
            status_text.text(f"Analyzing {tid}...")

            if tender["file"].endswith(".json"):
                path = os.path.join("../tenders", tender["file"])
                if os.path.exists(path):
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    content = f"""
            Tender Title: {data.get('title', '')}
            Issuer: {data.get('procuringEntity', {}).get('name', '')}
            Location: {data.get('procuringEntity', {}).get('address', {}).get('locality', '')}, {data.get('procuringEntity', {}).get('address', {}).get('region', '')}
            Budget: {data.get('value', {}).get('amount', '')} {data.get('value', {}).get('currency', 'UAH')}
            Deadline: {data.get('tenderPeriod', {}).get('endDate', '')}
            Description: {data.get('description', '')}
            """.strip()
            elif tender["file"].endswith(".pdf"):
                path = os.path.join("../extracted/", tender["file"].replace(".pdf", ".txt"))
                if os.path.exists(path):
                    content = extract_text_from_pdf(path)
                else:
                    st.warning(f"‚ö†Ô∏è PDF not found: {path}")
                    continue
            else:
                st.warning(f"Unsupported file type: {tender['file']}")
                continue


            result = analyze_tender(content, client)
            if result:
                result["tender_id"] = tid
                result["Filename"] = f"{tid}.txt"
                results.append(result)
            time.sleep(1.5)

        st.session_state.analysis_results = results
        status_text.text("‚úÖ –ê–Ω–∞–ª—ñ–∑ –∑–∞–≤–µ—Ä—à–µ–Ω–æ.")
        progress_bar.empty()

        # Save to Excel
        if results:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("–ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤")
            else:
                ws.title = "–ê–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤"
            ws, columns = format_excel(ws)
            for idx, res in enumerate(results, 2):
                row_data = [
                    res.get("title", "N/A"),
                    res.get("issuer", "N/A"),
                    res.get("deadline", "N/A"),
                    res.get("budget", "N/A"),
                    res.get("location", "N/A"),
                    res.get("project_type", "N/A"),
                    ", ".join(res.get("required_documents", [])),
                    "Yes" if res.get("avk5_required", False) else "No",
                    res.get("technical_specs", "N/A"),
                    res.get("payment_terms", "N/A"),
                    res.get("resource_requirements", "N/A"),
                    res.get("timeline_feasibility", "N/A"),
                    res.get("profitability", "N/A"),
                    res.get("Filename", "N/A")
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.session_state.excel_buffer = buf

    # üîÅ Always restore results
    results = st.session_state.get("analysis_results", [])
    if not results:
        st.warning("‚ö†Ô∏è No analysis results found. Run analysis first.")
        st.stop()

    # üìã SHOW ANALYSIS RESULTS
    st.subheader("–†–µ–∑—É–ª—å—Ç–∞—Ç–∏ –∞–Ω–∞–ª—ñ–∑—É")
    for res in results:
        with st.expander(f"{res.get('title', 'Untitled')} - {res.get('tender_id', '')}"):
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Issuer", res.get("issuer", "N/A"))
                st.metric("Deadline", res.get("deadline", "N/A"))
                st.metric("Budget", res.get("budget", "N/A"))
                st.metric("Location", res.get("location", "N/A"))
                st.metric("Project Type", res.get("project_type", "N/A"))
                st.metric("PC AVK5 Required", "‚úÖ Yes" if res.get("avk5_required") else "‚ùå No")
            with col2:
                st.subheader("Required Documents")
                for doc in res.get("required_documents", []):
                    st.write(f"- {doc}")
                st.subheader("Technical Specifications")
                st.info(res.get("technical_specs", "No technical specs"))
                st.subheader("Viability")
                st.metric("Timeline Feasibility", res.get("timeline_feasibility", "N/A"))
                st.metric("Profitability", res.get("profitability", "N/A"))

    if st.session_state.excel_buffer:
        st.download_button(
            label="üìä –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –ø–æ–≤–Ω–∏–π –∞–Ω–∞–ª—ñ–∑ (Excel)",
            data=st.session_state.excel_buffer,
            file_name="tender_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_download_button"
        )

elif tab == "üè¢ –ü—Ä–æ—Ñ—ñ–ª—å –∫–æ–º–ø–∞–Ω—ñ—ó":
    st.header("üè¢ –ü—Ä–æ—Ñ—ñ–ª—å –∫–æ–º–ø–∞–Ω—ñ—ó")
    
    st.subheader("–†–µ—Å—É—Ä—Å–∏ –∫–æ–º–ø–∞–Ω—ñ—ó")
    col1, col2, col3 = st.columns(3)
    workers = col1.number_input("–ö—ñ–ª—å–∫—ñ—Å—Ç—å —Ä–æ–±—ñ—Ç–Ω–∏–∫—ñ–≤", 
                               value=st.session_state.company_resources["workers"],
                               min_value=0, step=1)
    engineers = col2.number_input("–ö—ñ–ª—å–∫—ñ—Å—Ç—å —ñ–Ω–∂–µ–Ω–µ—Ä—ñ–≤", 
                                 value=st.session_state.company_resources["engineers"],
                                 min_value=0, step=1)
    vehicles = col3.number_input("–ö—ñ–ª—å–∫—ñ—Å—Ç—å —Ç—Ä–∞–Ω—Å–ø–æ—Ä—Ç–Ω–∏—Ö –∑–∞—Å–æ–±—ñ–≤", 
                                value=st.session_state.company_resources["vehicles"],
                                min_value=0, step=1)
    
    # Current projects
    st.subheader("–ü–æ—Ç–æ—á–Ω—ñ –ø—Ä–æ–µ–∫—Ç–∏")
    projects = st.session_state.company_resources["current_projects"]
    for i, project in enumerate(projects):
        col1, col2, col3 = st.columns([3, 2, 1])
        project_name = col1.text_input(f"–ù–∞–∑–≤–∞ –ø—Ä–æ–µ–∫—Ç—É {i+1}", value=project["name"])
        duration = col2.number_input(f"–¢—Ä–∏–≤–∞–ª—ñ—Å—Ç—å (–¥–Ω—ñ–≤)", value=project["duration"], min_value=1)
        if col3.button("‚ùå", key=f"del_proj_{i}"):
            projects.pop(i)
            st.rerun()
    
    if st.button("‚ûï –î–æ–¥–∞—Ç–∏ –ø—Ä–æ–µ–∫—Ç"):
        projects.append({"name": "–ù–æ–≤–∏–π –ø—Ä–æ–µ–∫—Ç", "duration": 30})
        st.rerun()
    
    # Document vault
    st.subheader("–°—Ö–æ–≤–∏—â–µ –¥–æ–∫—É–º–µ–Ω—Ç—ñ–≤")
    st.write("–ó–∞–≤–∞–Ω—Ç–∞–∂—Ç–µ –¥–æ–∫—É–º–µ–Ω—Ç–∏ –∫–æ–º–ø–∞–Ω—ñ—ó –¥–ª—è –ø–µ—Ä–µ–≤—ñ—Ä–∫–∏ –≤—ñ–¥–ø–æ–≤—ñ–¥–Ω–æ—Å—Ç—ñ")
    
    doc_name = st.text_input("–ù–∞–∑–≤–∞ –¥–æ–∫—É–º–µ–Ω—Ç–∞")
    doc_type = st.text_input("–¢–∏–ø –¥–æ–∫—É–º–µ–Ω—Ç–∞")
    validity = st.date_input("–î–∞—Ç–∞ –¥—ñ—ó")
    uploaded_file = st.file_uploader("–ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –¥–æ–∫—É–º–µ–Ω—Ç")
    
    if st.button("–î–æ–¥–∞—Ç–∏ –¥–æ —Å—Ö–æ–≤–∏—â–∞") and uploaded_file:
        # Save file
        os.makedirs("../data/documents", exist_ok=True)
        file_path = f"../data/documents/{uploaded_file.name}"
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        
        # Add to vault
        st.session_state.document_vault.add_document(
            doc_name, doc_type, validity.isoformat(), file_path
        )
        st.success(f"‚úÖ –î–æ–∫—É–º–µ–Ω—Ç '{doc_name}' –¥–æ–¥–∞–Ω–æ –¥–æ —Å—Ö–æ–≤–∏—â–∞!")
    
    # Save resources
    if st.button("üíæ –ó–±–µ—Ä–µ–≥—Ç–∏ –ø—Ä–æ—Ñ—ñ–ª—å –∫–æ–º–ø–∞–Ω—ñ—ó"):
        st.session_state.company_resources = {
            "workers": workers,
            "engineers": engineers,
            "vehicles": vehicles,
            "current_projects": projects
        }
        st.success("–ü—Ä–æ—Ñ—ñ–ª—å –∫–æ–º–ø–∞–Ω—ñ—ó –æ–Ω–æ–≤–ª–µ–Ω–æ!")

elif tab == "üìä –û—Ü—ñ–Ω–∫–∞ —Ç–µ–Ω–¥–µ—Ä–Ω–∏—Ö –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ–π":
    st.header("üìä –û—Ü—ñ–Ω–∫–∞ —Ç–µ–Ω–¥–µ—Ä–Ω–∏—Ö –ø—Ä–æ–ø–æ–∑–∏—Ü—ñ–π")

    if not st.session_state.analysis_results:
        st.warning("‚ö†Ô∏è –ù–µ –¥–æ—Å—Ç—É–ø–Ω–∏–π –∞–Ω–∞–ª—ñ–∑ —Ç–µ–Ω–¥–µ—Ä—ñ–≤. –ë—É–¥—å –ª–∞—Å–∫–∞, –∞–Ω–∞–ª—ñ–∑—É–π—Ç–µ —Ç–µ–Ω–¥–µ—Ä–∏ —Å–ø–æ—á–∞—Ç–∫—É.")
        st.stop()

    avk5_data = json.load(open("../data/avk5_standards.json", "r", encoding="utf-8"))
    compliance = st.session_state.document_vault
    profitability = ProfitabilityAnalyzer(AVK5Estimator())

    tender_options = {r["tender_id"]: r["title"] for r in st.session_state.analysis_results}
    selected_tender = st.selectbox("–í–∏–±–µ—Ä—ñ—Ç—å —Ç–µ–Ω–¥–µ—Ä –¥–ª—è –æ—Ü—ñ–Ω–∫–∏:", options=list(tender_options.keys()), format_func=lambda x: f"{tender_options[x][:50]}...")

    if not selected_tender:
        st.stop()

    tender_data = next(r for r in st.session_state.analysis_results if r["tender_id"] == selected_tender)

    # ---------------------- üìÑ Document Compliance ----------------------
    with st.expander("üìÑ –ü–µ—Ä–µ–≤—ñ—Ä–∫–∞ –≤—ñ–¥–ø–æ–≤—ñ–¥–Ω–æ—Å—Ç—ñ –¥–æ–∫—É–º–µ–Ω—Ç–∞", expanded=True):
        if tender_data.get("required_documents"):
            doc_report = compliance.check_compliance(tender_data["required_documents"])
            col1, col2 = st.columns([1, 3])
            col1.metric("–†—ñ–≤–µ–Ω—å –≤—ñ–¥–ø–æ–≤—ñ–¥–Ω–æ—Å—Ç—ñ", f"{doc_report['compliance_score']*100:.1f}%")
            col2.metric("–°—Ç–∞—Ç—É—Å", "‚úÖ –í—ñ–¥–ø–æ–≤—ñ–¥–∞—î" if doc_report["is_compliant"] else "‚ùå –ù–µ –≤—ñ–¥–ø–æ–≤—ñ–¥–∞—î")

            st.subheader("–°—Ç–∞—Ç—É—Å –¥–æ–∫—É–º–µ–Ω—Ç—ñ–≤")
            for doc in tender_data["required_documents"]:
                status = "‚úÖ –î–æ—Å—Ç—É–ø–Ω–∏–π" if doc in doc_report["available_documents"] else "‚ùå –í—ñ–¥—Å—É—Ç–Ω—ñ–π"
                st.write(f"{status} - {doc}")
        else:
            st.info("–ù–µ–º–∞—î –¥–æ–∫—É–º–µ–Ω—Ç—ñ–≤, —â–æ –≤–∫–∞–∑–∞–Ω—ñ –≤ —Ç–µ–Ω–¥–µ—Ä—ñ.")

    # ---------------------- üß± AVK5 Material Cost Estimation ----------------------
    st.subheader("üß± –î–æ–¥–∞–π—Ç–µ –≤–ª–∞—Å–Ω—ñ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏ –¥–ª—è –æ—Ü—ñ–Ω–∫–∏ AVK5")

    # Prepare dropdown materials
    material_options = []
    category_map = {}
    for cat, items in avk5_data.items():
        if isinstance(items, dict):
            for spec in items:
                material_options.append(f"{spec} ({cat})")
                category_map[spec] = cat

    if "custom_materials" not in st.session_state:
        st.session_state.custom_materials = []

    selected_material = st.selectbox("–í–∏–±–µ—Ä—ñ—Ç—å –º–∞—Ç–µ—Ä—ñ–∞–ª", material_options)
    mat_spec = selected_material.split(" (")[0]
    mat_cat = category_map[mat_spec]
    unit = avk5_data[mat_cat][mat_spec]["unit"]
    default_price = avk5_data[mat_cat][mat_spec]["price"]

    with st.form("material_form"):
        col1, col2, col3 = st.columns(3)
        qty = col1.number_input("–ö—ñ–ª—å–∫—ñ—Å—Ç—å", min_value=0.0, step=0.1, key="mat_qty")
        price = col2.number_input("–¶—ñ–Ω–∞ –∑–∞ –æ–¥–∏–Ω–∏—Ü—é (UAH)", value=float(default_price), step=100.0, key="mat_price")
        col3.markdown(f"üí° –û–¥–∏–Ω–∏—Ü—è –≤–∏–º—ñ—Ä—É: **{unit}**")
        if st.form_submit_button("‚ûï –î–æ–¥–∞—Ç–∏ –º–∞—Ç–µ—Ä—ñ–∞–ª"):
            if qty > 0:
                st.session_state.custom_materials.append({
                    "category": mat_cat,
                    "specification": mat_spec,
                    "quantity": qty,
                    "unit_price": price,
                    "total": qty * price
                })
                st.success(f"‚úÖ {mat_spec} –¥–æ–¥–∞–Ω–æ!")

    if st.session_state.custom_materials:
        st.subheader("üì¶ –ü–æ—Ç–æ—á–Ω—ñ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏")
        total = 0
        for m in st.session_state.custom_materials:
            st.write(f"- {m['specification']} ({m['category']}): {m['quantity']} √ó {m['unit_price']} = {m['total']:.2f} UAH")
            total += m["total"]
        st.markdown(f"### üí∞ –ó–∞–≥–∞–ª—å–Ω–∞ –≤–∞—Ä—Ç—ñ—Å—Ç—å –º–∞—Ç–µ—Ä—ñ–∞–ª—ñ–≤: **{total:,.2f} UAH**")

        # Export to Excel
        if st.button("üì• –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –≤–∞—Ä—Ç—ñ—Å—Ç—å –º–∞—Ç–µ—Ä—ñ–∞–ª—ñ–≤ (Excel)"):
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="–í–ª–∞—Å–Ω—ñ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏")
            else:
                ws.title = "–í–ª–∞—Å–Ω—ñ –º–∞—Ç–µ—Ä—ñ–∞–ª–∏"
            headers = ["Category", "Specification", "Quantity", "Unit Price", "Total"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            for i, m in enumerate(st.session_state.custom_materials, 2):
                ws.cell(row=i, column=1, value=m["category"])
                ws.cell(row=i, column=2, value=m["specification"])
                ws.cell(row=i, column=3, value=m["quantity"])
                ws.cell(row=i, column=4, value=m["unit_price"])
                ws.cell(row=i, column=5, value=m["total"])
            ws.cell(row=len(st.session_state.custom_materials) + 2, column=4, value="Total:")
            ws.cell(row=len(st.session_state.custom_materials) + 2, column=5, value=total)

            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            st.download_button("üì• –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –æ—Ü—ñ–Ω–∫—É AVK5", data=excel_buffer, file_name=f"avk5_materials_{selected_tender}.xlsx")

    # ---------------------- üí∞ Profitability Analysis ----------------------
    with st.expander("üí∞ –ê–Ω–∞–ª—ñ–∑ –ø—Ä–∏–±—É—Ç–∫–æ–≤–æ—Å—Ç—ñ", expanded=True):
        def extract_resources(text):
            import re
            resources = {"workers": 0, "engineers": 0, "vehicles": 0}
            patterns = {
                "workers": r"(\d+)\s*(workers|laborers|people)",
                "engineers": r"(\d+)\s*(engineers)",
                "vehicles": r"(\d+)\s*(vehicles|trucks|cars)"
            }
            for key, pattern in patterns.items():
                found = re.search(pattern, text, re.IGNORECASE)
                if found:
                    resources[key] = int(found.group(1))
            return resources

        def estimate_complexity(text):
            text = text.lower()
            if any(w in text for w in ["automation", "bim", "hvac", "deep foundation"]): return 8
            if any(w in text for w in ["roof", "paving", "electrical"]): return 5
            if any(w in text for w in ["painting", "doors"]): return 3
            return 4

        def safe_budget(budget_raw):
            try:
                return float(budget_raw.split()[0].replace(",", ""))
            except:
                return 0.0

        auto_resource_req = extract_resources(tender_data.get("resource_requirements", ""))
        auto_complexity = estimate_complexity(tender_data.get("technical_specs", ""))
        auto_budget = safe_budget(tender_data.get("budget", "0"))
        # Calculate total cost from AVK5 custom materials if available
        custom_materials = st.session_state.get("custom_materials", [])
        estimated_cost = 0

        if custom_materials:
            estimated_cost = sum(m["total"] for m in custom_materials)
            st.markdown(f"üí∏ **–û—Ü—ñ–Ω–µ–Ω–∞ –≤–∞—Ä—Ç—ñ—Å—Ç—å –∑ –≤–ª–∞—Å–Ω–∏—Ö –º–∞—Ç–µ—Ä—ñ–∞–ª—ñ–≤**: `{estimated_cost:,.2f} UAH`")

        tender = {
            "title": tender_data.get("title", ""),
            "budget": auto_budget,
            "resource_requirements": auto_resource_req,
            "estimated_cost": estimated_cost,
            "timeline": {
                "duration_days": 90,
                "start_date": "2025-09-01"
            },
            "complexity": auto_complexity,
            "payment_terms": tender_data.get("payment_terms", "standard").lower(),
            "has_penalties": False,
            "competitors": 3,
            "required_docs": tender_data.get("required_documents", [])
        }

        company = st.session_state.company_resources
        analysis = profitability.analyze_tender(tender, company)

        col1, col2, col3 = st.columns(3)
        col1.metric("–†—ñ–≤–µ–Ω—å ROI", f"{analysis['roi_score']:.1f}/100")
        col2.metric("–†—ñ–≤–µ–Ω—å –ø—Ä–∏–±—É—Ç–∫–æ–≤–æ—Å—Ç—ñ", f"{analysis['profit_margin']*100:.1f}%")
        col3.metric("–†–µ–∫–æ–º–µ–Ω–¥–∞—Ü—ñ—è", analysis["recommendation"])

        st.subheader("–ê–Ω–∞–ª—ñ–∑ –¥–µ—Ñ—ñ—Ü–∏—Ç—É —Ä–µ—Å—É—Ä—Å—ñ–≤")
        st.subheader("üìà –§—ñ–Ω–∞–Ω—Å–æ–≤–∏–π –∞–Ω–∞–ª—ñ–∑")
        st.markdown(f"- **–í–∞—Ä—Ç—ñ—Å—Ç—å —Ç–µ–Ω–¥–µ—Ä–∞**: {analysis['tender_value']:,.2f} UAH")
        st.markdown(f"- **–û—Ü—ñ–Ω–µ–Ω–∞ –≤–∞—Ä—Ç—ñ—Å—Ç—å**: {analysis['estimated_cost']:,.2f} UAH")
        st.markdown(f"- **–í–∞–ª–æ–≤–∏–π –ø—Ä–∏–±—É—Ç–æ–∫**: {analysis['gross_profit']:,.2f} UAH")

        st.subheader("üì¶ –ê–Ω–∞–ª—ñ–∑ –≤–∏—Ç—Ä–∞—Ç")
        for cat, val in analysis.get("cost_breakdown", {}).items():
            if isinstance(val, (int, float)):
                st.write(f"- {cat}: {val:,.2f} UAH")
            elif isinstance(val, list):
                st.write(f"- {cat}:")
                for item in val:
                    st.write(f"    {item}")
            else:
                st.write(f"- {cat}: {val}")

        st.subheader("‚è±Ô∏è –ê–Ω–∞–ª—ñ–∑ –º–æ–∂–ª–∏–≤–æ—Å—Ç—ñ –≤–∏–∫–æ–Ω–∞–Ω–Ω—è")
        st.info(f"{analysis.get('timeline_feasibility', 'N/A')}")

        st.subheader("‚ö†Ô∏è –†–∏–∑–∏–∫–∏")
        for factor in analysis.get("risk_factors", []):
            st.warning(f"- {factor}")

        if analysis["resource_gap"].get("gap_analysis"):
            for resource, data in analysis["resource_gap"]["gap_analysis"].items():
                gap_percent = max(0, min(1, data["gap_percent"]))
                st.progress(1 - gap_percent, text=f"{resource}: {data['available']}/{data['required']} ({data['gap']} gap)")
        else:
            st.info("–ù–µ–º–∞—î —Ä–µ—Å—É—Ä—Å–Ω–∏—Ö –≤–∏–º–æ–≥.")

        # Export Evaluation to Excel
        if st.button("üì• –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –∑–≤—ñ—Ç –ø—Ä–æ –æ—Ü—ñ–Ω–∫—É (Excel)"):
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet(title="–û—Ü—ñ–Ω–∫–∞ —Ç–µ–Ω–¥–µ—Ä–∞")
            else:
                ws.title = "–û—Ü—ñ–Ω–∫–∞ —Ç–µ–Ω–¥–µ—Ä–∞"
                
            ws.append(["–ù–∞–∑–≤–∞ —Ç–µ–Ω–¥–µ—Ä–∞", tender["title"]])
            ws.append(["–ë—é–¥–∂–µ—Ç", tender["budget"]])
            ws.append(["–¢—Ä–∏–≤–∞–ª—ñ—Å—Ç—å (–¥–Ω—ñ–≤)", tender["timeline"]["duration_days"]])
            ws.append(["–°–∫–ª–∞–¥–Ω—ñ—Å—Ç—å", tender["complexity"]])
            ws.append(["–£–º–æ–≤–∏ –æ–ø–ª–∞—Ç–∏", tender["payment_terms"]])
            ws.append(["–ù–∞—è–≤–Ω—ñ—Å—Ç—å —à—Ç—Ä–∞—Ñ—ñ–≤", tender["has_penalties"]])
            ws.append(["–ö–æ–Ω–∫—É—Ä–µ–Ω—Ç–∏", tender["competitors"]])
            ws.append(["–†—ñ–≤–µ–Ω—å ROI", analysis["roi_score"]])
            ws.append(["–†—ñ–≤–µ–Ω—å –ø—Ä–∏–±—É—Ç–∫–æ–≤–æ—Å—Ç—ñ", analysis["profit_margin"]])
            ws.append(["–†–µ–∫–æ–º–µ–Ω–¥–∞—Ü—ñ—è", analysis["recommendation"]])
            ws.append([])

            # Financial Breakdown
            ws.append(["üìà –§—ñ–Ω–∞–Ω—Å–æ–≤–∏–π –∞–Ω–∞–ª—ñ–∑"])
            ws.append(["–í–∞—Ä—Ç—ñ—Å—Ç—å —Ç–µ–Ω–¥–µ—Ä–∞", analysis["tender_value"]])
            ws.append(["–û—Ü—ñ–Ω–µ–Ω–∞ –≤–∞—Ä—Ç—ñ—Å—Ç—å", analysis["estimated_cost"]])
            ws.append(["–í–∞–ª–æ–≤–∏–π –ø—Ä–∏–±—É—Ç–æ–∫", analysis["gross_profit"]])
            ws.append([])

            # Cost Breakdown
            ws.append(["üì¶ –ê–Ω–∞–ª—ñ–∑ –≤–∏—Ç—Ä–∞—Ç"])
            ws.append(["–ö–∞—Ç–µ–≥–æ—Ä—ñ—è", "–°—É–º–∞ (UAH)"])
            for cat, val in analysis.get("cost_breakdown", {}).items():
                ws.append([cat, val])
            ws.append([])

            # Timeline Feasibility
            ws.append(["‚è±Ô∏è –ê–Ω–∞–ª—ñ–∑ –º–æ–∂–ª–∏–≤–æ—Å—Ç—ñ –≤–∏–∫–æ–Ω–∞–Ω–Ω—è", analysis.get("timeline_feasibility", "N/A")])
            ws.append([])

            # Risk Factors
            ws.append(["‚ö†Ô∏è –†–∏–∑–∏–∫–∏"])
            for risk in analysis.get("risk_factors", []):
                ws.append([risk])
            ws.append([])

            # Resource Gaps
            ws.append(["–†–µ—Å—É—Ä—Å", "–ü–æ—Ç—Ä—ñ–±–Ω–æ", "–ù–∞—è–≤–Ω—ñ—Å—Ç—å", "–î–µ—Ñ—ñ—Ü–∏—Ç", "–†—ñ–≤–µ–Ω—å –¥–µ—Ñ—ñ—Ü–∏—Ç—É"])
            for resource, data in analysis["resource_gap"]["gap_analysis"].items():
                ws.append([
                    resource,
                    data["required"],
                    data["available"],
                    data["gap"],
                    f"{data['gap_percent']*100:.1f}%"
                ])

            # Export
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="üìä –ó–∞–≤–∞–Ω—Ç–∞–∂–∏—Ç–∏ –∑–≤—ñ—Ç –ø—Ä–æ –æ—Ü—ñ–Ω–∫—É (Excel)",
                data=buffer,
                file_name=f"tender_evaluation_{selected_tender}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )