import streamlit as st
from datetime import datetime, timedelta
import json
import os
import sys
import time
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from io import BytesIO
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(script_dir, '..'))
from core.score_matrix import AVK5Estimator, DocumentComplianceChecker, ProfitabilityAnalyzer
from core.company_profile import CompanyProfile

# Add core modules to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from core.downloader import download_prozorro_tenders, setup_environment

# Load environment variables
load_dotenv()
def build_tender_text(tender_json: dict) -> str:
    title = tender_json.get("title", "")
    description = tender_json.get("description", "")
    issuer = tender_json.get("procuringEntity", {}).get("name", "")
    address = tender_json.get("procuringEntity", {}).get("address", {})
    location = f"{address.get('locality', '')}, {address.get('region', '')}"
    budget = tender_json.get("value", {}).get("amount", "N/A")
    currency = tender_json.get("value", {}).get("currency", "UAH")
    deadline = tender_json.get("tenderPeriod", {}).get("endDate", "Not specified")

    # Extract technical requirements (if available)
    tech_specs = []
    for criterion in tender_json.get("criteria", []):
        for group in criterion.get("requirementGroups", []):
            for req in group.get("requirements", []):
                title = req.get("title", "")
                expected = req.get("expectedValues", []) or [req.get("expectedValue", "")]
                tech_specs.append(f"{title}: {', '.join(str(v) for v in expected if v)}")

    specs_block = "\n".join(tech_specs)

    return f"""
Tender Title: {title}
Issuer: {issuer}
Location: {location}
Budget: {budget} {currency}
Deadline: {deadline}
Description: {description}

Technical Requirements:
{specs_block}
""".strip()

# Initialize Claude client
def get_claude_client():
    api_key = os.getenv("CLAUDE_API_KEY")
    if not api_key:
        st.error("❌ Claude API key not found in .env file")
        return None
    return anthropic.Anthropic(api_key=api_key)

# Enhanced parser function
def analyze_tender(text, client):
    prompt = f"""
You are an expert in Ukrainian public procurement tenders. Analyze the tender text and extract the following information:

... (same as before)

Tender Text:
\"\"\"
{text[:15000]}
\"\"\"
"""
    try:
        st.write("🔎 Prompt preview:", prompt[:1000])  # Log first 1000 characters
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1024,
            temperature=0.0,
            system="You are a procurement specialist analyzing Ukrainian tenders. Focus on PC AVK5 compliance and document requirements.",
            messages=[{"role": "user", "content": prompt}]
        )
        
        if response.content:
            result_text = ''.join(block.text for block in response.content if hasattr(block, 'text'))
            st.write("🧾 Raw Claude output:", result_text[:1000])  # First 1000 chars
            try:
                return json.loads(result_text)
            except json.JSONDecodeError:
                import re
                match = re.search(r'\{[\s\S]+\}', result_text)
                if match:
                    try:
                        return json.loads(match.group())
                    except Exception as e:
                        st.error(f"❌ Failed to parse Claude output: {e}")
        else:
            st.warning("⚠️ Claude returned empty content.")
    except Exception as e:
        st.error(f"Claude API error: {str(e)}")
    return {}
# Excel formatting
def format_excel(ws):
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    col_widths = [40, 30, 15, 15, 20, 25, 40, 15, 50, 30, 40, 20, 20, 30]
    columns = [
        "Title", "Issuer", "Deadline", "Budget", "Location", 
        "Project Type", "Required Documents", "PC AVK5 Required",
        "Technical Specifications", "Payment Terms", "Resource Requirements",
        "Timeline Feasibility", "Profitability Assessment", "Filename"
    ]
    
    for col_num, (column_title, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    return ws, columns

# Streamlit App
st.set_page_config(page_title="AI Tender Optimizer", layout="wide")

# Initialize session state
if "analysis_results" not in st.session_state:
    st.session_state.analysis_results = []
if "tenders_downloaded" not in st.session_state:
    st.session_state.tenders_downloaded = []
if "excel_buffer" not in st.session_state:
    st.session_state.excel_buffer = None
if "analysis_attempted" not in st.session_state:
    st.session_state.analysis_attempted = False

# Sidebar
st.sidebar.title("🛠️ Tender Settings")
tab = st.sidebar.radio("Navigation", ["📥 Data Downloader", "🔍 Tender Analysis"])

keywords_option = st.sidebar.multiselect(
    "🔍 Select Tender Topics",
    ["IT", "Construction", "Medical Equipment", "Transport", "Education"],
    default=["Construction"]
)

days_range = st.sidebar.slider(
    "📅 Days Back to Search",
    min_value=1,
    max_value=60,
    value=30
)

# Main Page
st.title("📦Tender Optimizer")

if tab == "📥 Data Downloader":
    run_download = st.button("🚀 Download Tenders", key="download_button")
    
    if run_download:
        st.info("⏳ Starting download process...")
        setup_environment()

        keyword_map = {
            "IT": "IT",
            "Construction": "будівництво",
            "Medical Equipment": "медичне обладнання",
            "Transport": "транспорт",
            "Education": "освіта"
        }
        selected_keywords = [keyword_map[k] for k in keywords_option]

        with st.spinner("Downloading tenders from Prozorro..."):
            tenders = download_prozorro_tenders(
                keywords=selected_keywords,
                days_back=days_range
            )

        if tenders:
            st.session_state.tenders_downloaded = tenders
            st.success(f"✅ {len(tenders)} tenders downloaded and saved locally.")
            with st.expander("📄 View Summary"):
                st.json({
                    "total": len(tenders),
                    "keywords": selected_keywords,
                    "date_range": f"Last {days_range} days"
                })

            st.download_button(
                label="📥 Download Summary JSON",
                data=json.dumps(tenders, ensure_ascii=False, indent=2),
                file_name="prozorro_tenders_summary.json",
                mime="application/json"
            )
        else:
            st.warning("⚠️ No tenders were downloaded or an error occurred.")

elif tab == "🔍 Tender Analysis":
    st.header("🔍 Tender Analysis")
    
    if not st.session_state.tenders_downloaded:
        st.warning("⚠️ No tenders downloaded yet. Please download tenders first.")
        st.stop()
        
    # Tender selection
    tender_options = {t['id']: t['title'] for t in st.session_state.tenders_downloaded}
    selected_tenders = st.multiselect(
        "Select tenders to analyze:",
        options=list(tender_options.keys()),
        format_func=lambda x: f"{x} - {tender_options[x][:50]}..."
    )
    
    if not selected_tenders:
        st.info("ℹ️ Please select at least one tender to analyze")
        st.stop()
        
    # Get Claude client
    client = get_claude_client()
    if not client:
        st.stop()
    
    # Analysis button with unique key
    analyze_clicked = st.button("🔍 Analyze Selected Tenders", key="analyze_button")
    
    if analyze_clicked:
        st.session_state.analysis_attempted = True
        analysis_results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, tender_id in enumerate(selected_tenders):
            # Find the tender data
            tender_data = next((t for t in st.session_state.tenders_downloaded 
                               if t['id'] == tender_id), None)
            if not tender_data:
                continue
                
            # Update progress
            progress = (i + 1) / len(selected_tenders)
            progress_bar.progress(progress)
            status_text.text(f"Analyzing tender {i+1}/{len(selected_tenders)}: {tender_id}")
            
            # Read tender text
            json_path = os.path.join("tenders", f"ProZorro_{tender_id}.json")
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    tender_json = json.load(f)

                # Compose a tender text for Claude
                tender_text = build_tender_text(tender_json)
                
                # Analyze with Claude
                result = analyze_tender(tender_text, client)
                if result:
                    result["tender_id"] = tender_id
                    result["Filename"] = f"ProZorro_{tender_id}.json"
                    analysis_results.append(result)
            
            # Sleep to avoid rate limits
            time.sleep(1.5)
        
        status_text.text("✅ Analysis complete!")
        progress_bar.empty()
        
        # Save results to session state
        st.session_state.analysis_results = analysis_results
        
        # Create Excel workbook
        if analysis_results:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("Tender Analysis")
            else:
                ws.title = "Tender Analysis"
            ws, columns = format_excel(ws)
            
            # Add data to Excel
            for row_idx, result in enumerate(analysis_results, 2):
                row_data = [
                    result.get("title", "N/A"),
                    result.get("issuer", "N/A"),
                    result.get("deadline", "N/A"),
                    result.get("budget", "N/A"),
                    result.get("location", "N/A"),
                    result.get("project_type", "N/A"),
                    ", ".join(result.get("required_documents", [])),
                    "Yes" if result.get("avk5_required", False) else "No",
                    result.get("technical_specs", "N/A"),
                    result.get("payment_terms", "N/A"),
                    result.get("resource_requirements", "N/A"),
                    result.get("timeline_feasibility", "N/A"),
                    result.get("profitability", "N/A"),
                    result.get("Filename", "N/A")
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Save to bytes for download
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            st.session_state.excel_buffer = excel_buffer
    
    # Display results if available
    if st.session_state.analysis_results:
        st.subheader("Analysis Results")
        
        for result in st.session_state.analysis_results:
            with st.expander(f"{result.get('title', 'Untitled Tender')} - {result.get('tender_id', '')}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.subheader("Basic Information")
                    st.metric("Issuer", result.get("issuer", "N/A"))
                    st.metric("Deadline", result.get("deadline", "N/A"))
                    st.metric("Budget", result.get("budget", "N/A"))
                    st.metric("Location", result.get("location", "N/A"))
                    st.metric("Project Type", result.get("project_type", "N/A"))
                    
                    st.subheader("Compliance")
                    st.metric("PC AVK5 Required", 
                             "✅ Yes" if result.get("avk5_required") else "❌ No")
                
                with col2:
                    st.subheader("Requirements")
                    st.write("**Required Documents:**")
                    docs = result.get("required_documents", [])
                    if docs:
                        for doc in docs:
                            st.write(f"- {doc}")
                    else:
                        st.write("No documents listed")
                    
                    st.subheader("Technical Specifications")
                    st.info(result.get("technical_specs", "No technical specs available"))
                    
                    st.subheader("Viability")
                    col_f, col_p = st.columns(2)
                    col_f.metric("Timeline Feasibility", result.get("timeline_feasibility", "N/A"))
                    col_p.metric("Profitability", result.get("profitability", "N/A"))
        
        # Download button with unique key
        if st.session_state.excel_buffer:
            st.download_button(
                label="📊 Download Full Analysis (Excel)",
                data=st.session_state.excel_buffer,
                file_name="tender_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="excel_download_button"
            )
    
    # Show warning if analysis was attempted but no results
    elif st.session_state.analysis_attempted and not st.session_state.analysis_results:
        st.warning("No analysis results were generated. Please check the tender files and try again.")
        # Add to your Streamlit app
elif tab == "📊 Tender Evaluation":
    st.header("📊 Comprehensive Tender Evaluation")
    
    if not st.session_state.analysis_results:
        st.warning("⚠️ No tender analysis available. Please analyze tenders first.")
        st.stop()
        
    # Select tender to evaluate
    tender_options = {r["tender_id"]: r["title"] for r in st.session_state.analysis_results}
    selected_tender = st.selectbox(
        "Select tender for detailed evaluation:",
        options=list(tender_options.keys()),
        format_func=lambda x: f"{x} - {tender_options[x][:50]}..."
    )
    
    if not selected_tender:
        st.stop()
        
    tender_data = next(r for r in st.session_state.analysis_results if r["tender_id"] == selected_tender)
    
    # Initialize modules
    avk5 = AVK5Estimator()
    compliance = DocumentComplianceChecker()
    profitability = ProfitabilityAnalyzer(avk5)
    
    # Display evaluation sections
    with st.expander("📄 Document Compliance Check"):
        doc_report = compliance.check_compliance(tender_data.get("required_documents", []))
        
        col1, col2 = st.columns([1, 3])
        col1.metric("Compliance Score", f"{doc_report['compliance_score']*100:.1f}%")
        col2.metric("Status", "✅ Compliant" if doc_report["is_compliant"] else "❌ Non-Compliant")
        
        st.subheader("Document Status")
        for doc in tender_data.get("required_documents", []):
            status = "✅ Available" if doc in doc_report["available_documents"] else "❌ Missing"
            st.write(f"{status} - {doc}")
            
        if not doc_report["is_compliant"]:
            st.warning("Missing documents detected!")
            for doc in doc_report["missing_documents"]:
                st.write(f"❌ {doc} - Not in company vault")
    
    with st.expander("🧮 PC AVK5 Cost Estimation"):
        if tender_data.get("avk5_required", "No").lower() == "yes":
            # Sample construction data - in real app, extract from tender docs
            construction_data = {
                "materials": {"concrete": (120, "M300"), "rebar": (8, "A500C-Ø12")},
                "labor": {"mason": (240, "standard"), "carpenter": (120, "standard")},
                "equipment": {"crane_25t": (1, 5), "concrete_pump": (1, 3)}
            }
            
            estimate = avk5.calculate_estimate(
                construction_data["materials"],
                construction_data["labor"],
                construction_data["equipment"]
            )
            
            st.metric("Final Price", f"{estimate['final_price']:,.2f} UAH")
            st.download_button(
                "📥 Download AVK5 Estimate",
                data=avk5.export_to_excel(estimate, "avk5_estimate.xlsx"),
                file_name=f"avk5_estimate_{selected_tender}.xlsx"
            )
        else:
            st.info("This tender does not require PC AVK5 cost estimation")
    
    with st.expander("💰 Profitability Analysis"):
        # Sample company data - should come from company profile
        company_resources = {
            "workers": 12,
            "engineers": 3,
            "vehicles": 2,
            "current_projects": [
                {"name": "Hospital Project", "duration": 45},
                {"name": "Apartment Building", "duration": 30}
            ]
        }
        
        analysis = profitability.analyze_tender(tender_data, company_resources)
        
        col1, col2, col3 = st.columns(3)
        col1.metric("ROI Score", f"{analysis['roi_score']:.1f}/100")
        col2.metric("Profit Margin", f"{analysis['profit_margin']*100:.1f}%")
        col3.metric("Recommendation", analysis['recommendation'], 
                   delta_color="inverse" if "NO-BID" in analysis['recommendation'] else "normal")
        
        st.subheader("Resource Gap Analysis")
        for resource, data in analysis["resource_gap"]["gap_analysis"].items():
            st.progress(1 - data["gap_percent"], 
                        text=f"{resource}: {data['available']}/{data['required']}")
        # Add to Streamlit app
elif tab == "🏢 Company Profile":
    st.header("🏢 Company Profile Management")
    
    
    if "company_profile" not in st.session_state:
       st.session_state.company_profile = CompanyProfile()
    profile = st.session_state.company_profile  # This is the instance
    
    with st.expander("📁 Document Vault"):
        st.subheader("Upload New Document")
        doc_name = st.text_input("Document Name")
        doc_type = st.text_input("Document Type")
        validity = st.date_input("Validity Date")
        uploaded_file = st.file_uploader("Upload Document")
        
        if st.button("Add to Vault") and uploaded_file:
            # Save file
            file_path = f"docs/{uploaded_file.name}"
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Add to profile
            profile.add_document(
                doc_name, doc_type, validity.isoformat(), file_path
            )
            st.success("Document added to vault!")
        
        st.subheader("Existing Documents")
        for doc in profile.profile["document_vault"]:
            col1, col2 = st.columns([3, 1])
            col1.write(f"**{doc['name']}** ({doc['type']})")
            col1.caption(f"Valid until: {doc['validity']}")
            col2.download_button(
                "Download", 
                data=open(doc["path"], "rb").read(),
                file_name=doc["path"].split("/")[-1],
                key=f"dl_{doc['id']}"
            )
    
    with st.expander("🛠️ Resource Capabilities"):
        st.subheader("Company Resources")
        resources = profile.profile["resources"]
        workers = st.number_input("Number of Workers", value=resources.get("workers", 0))
        engineers = st.number_input("Number of Engineers", value=resources.get("engineers", 0))
        vehicles = st.number_input("Number of Vehicles", value=resources.get("vehicles", 0))
        
        if st.button("Update Resources"):
            profile.update_resources({
                "workers": workers,
                "engineers": engineers,
                "vehicles": vehicles
            })
            st.success("Resources updated!")
    
    with st.expander("📈 Historical Performance"):
        if st.button("Add New Performance Record"):
            with st.form("performance_form"):
                tender_id = st.text_input("Tender ID")
                outcome = st.selectbox("Outcome", ["Won", "Lost", "Cancelled"])
                profit = st.number_input("Profit (UAH)", value=0.0)
                lessons = st.text_area("Lessons Learned")
                
                if st.form_submit_button("Save Record"):
                    profile.add_performance_record(tender_id, outcome, profit, lessons)
        
        for record in profile.profile["historical_performance"]:
            st.write(f"**Tender {record['tender_id']}** - {record['outcome']}")
            st.write(f"Profit: {record['profit']} UAH")
            st.caption(f"Lessons: {record['lessons']}")